"""
Servicio de comparación entre BD y archivos Excel externos.
Usa Pandas para el análisis de diferencias.
"""
import pandas as pd
import re
import time
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple, Set
from sqlalchemy.orm import Session

from app.utils.logger import app_logger as logger
from app.utils.validators import validador


class ComparacionService:
    """
    Compara registros de la BD con un archivo Excel externo.
    
    Claves de comparación: numero_identificacion
    Campos comparados: nombres, apellidos, fecha_nacimiento, 
                       fecha_expedicion, lugar_expedicion, sexo
    
    Tipos de diferencias:
    - igual: mismo ID, mismos datos
    - diferente: mismo ID, datos distintos en algún campo
    - faltante_bd: está en Excel pero no en BD
    - nuevo_bd: está en BD pero no en Excel
    """

    CAMPOS_COMPARACION = [
        "nombres",
        "apellidos",
        "fecha_nacimiento",
        "fecha_expedicion",
        "lugar_expedicion",
        "sexo",
    ]

    MAPEO_COLUMNAS_EXCEL = {
        # Identificación
        "identificacion": "numero_identificacion",
        "numero_identificacion": "numero_identificacion",
        "numero_de_identificacion": "numero_identificacion",
        "num_identificacion": "numero_identificacion",
        "nro_identificacion": "numero_identificacion",
        "cedula": "numero_identificacion",
        "cedula_de_ciudadania": "numero_identificacion",
        "cc": "numero_identificacion",
        "c_c": "numero_identificacion",
        "documento": "numero_identificacion",
        "numero_documento": "numero_identificacion",
        "numero_de_documento": "numero_identificacion",
        "num_documento": "numero_identificacion",
        "no_documento": "numero_identificacion",
        "nro_documento": "numero_identificacion",
        "documento_de_identidad": "numero_identificacion",
        "documento_identidad": "numero_identificacion",
        "doc": "numero_identificacion",
        "nuip": "numero_identificacion",
        "ti": "numero_identificacion",
        "tarjeta_identidad": "numero_identificacion",
        "id": "numero_identificacion",

        # Nombres
        "nombre": "nombres",
        "nombres": "nombres",
        "primer_nombre": "primer_nombre",
        "segundo_nombre": "segundo_nombre",
        "nombre_completo": "nombres",

        # Apellidos
        "apellido": "apellidos",
        "apellidos": "apellidos",
        "primer_apellido": "primer_apellido",
        "segundo_apellido": "segundo_apellido",

        # Fechas
        "fecha_nacimiento": "fecha_nacimiento",
        "fecha_de_nacimiento": "fecha_nacimiento",
        "f_nacimiento": "fecha_nacimiento",
        "f_nac": "fecha_nacimiento",
        "nacimiento": "fecha_nacimiento",

        "fecha_expedicion": "fecha_expedicion",
        "fecha_de_expedicion": "fecha_expedicion",
        "f_expedicion": "fecha_expedicion",
        "f_exp": "fecha_expedicion",
        "expedicion": "fecha_expedicion",

        # Lugar
        "lugar_expedicion": "lugar_expedicion",
        "lugar_de_expedicion": "lugar_expedicion",
        "lugar": "lugar_expedicion",
        "ciudad_expedicion": "lugar_expedicion",
        "ciudad_de_expedicion": "lugar_expedicion",
        "municipio_expedicion": "lugar_expedicion",
        "municipio_de_expedicion": "lugar_expedicion",
        "ciudad": "lugar_expedicion",
        "municipio": "lugar_expedicion",

        # Sexo / Género
        "sexo": "sexo",
        "genero": "sexo",
        "sex": "sexo",
    }

    def _normalizar_nombre_columna(self, col: Any) -> str:
        """Normaliza un nombre de columna: minúsculas, sin tildes, sin signos."""
        s = str(col).lower().strip()
        repl = {
            "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n",
            ".": "", "-": " ", "/": " ", "\\": " ", "_": " ", "°": "", "#": ""
        }
        for k, v in repl.items():
            s = s.replace(k, v)
        s = re.sub(r"\s+", "_", s).strip("_")
        return s

    def _limpiar_numero_id(self, val: Any) -> str:
        """Limpia número de identificación: quita puntos, comas, espacios y decimales .0 de Excel."""
        if val is None or pd.isna(val):
            return ""
        s = str(val).strip()
        if s.endswith(".0"):
            s = s[:-2]
        s = re.sub(r"[^\d]", "", s)
        return s
    def _son_nombres_equivalentes(
        self,
        nombres_bd: str,
        apellidos_bd: str,
        nombres_excel: str,
        apellidos_excel: str = ""
    ) -> bool:
        """
        Verifica si los nombres/apellidos coinciden independientemente de si el Excel
        los tiene en una sola columna ('JEISON BASTIDAS ORTIZ') o separados.
        """
        norm_nb = self._normalizar_para_comparacion(nombres_bd)
        norm_ab = self._normalizar_para_comparacion(apellidos_bd)
        norm_ne = self._normalizar_para_comparacion(nombres_excel)
        norm_ae = self._normalizar_para_comparacion(apellidos_excel)

        # Si el Excel tiene nombres y apellidos en columnas separadas
        if norm_ae:
            if norm_nb == norm_ne and norm_ab == norm_ae:
                return True

        # Si el Excel tiene todo en una sola columna (Nombre Completo)
        palabras_bd = set(re.findall(r"[A-Z0-9]+", f"{norm_nb} {norm_ab}"))
        palabras_excel = set(re.findall(r"[A-Z0-9]+", f"{norm_ne} {norm_ae}"))

        if not palabras_bd or not palabras_excel:
            return False

        # Si el conjunto de palabras es idéntico
        inter = palabras_bd.intersection(palabras_excel)
        if len(inter) == len(palabras_excel) and len(inter) == len(palabras_bd):
            return True

        # Coincidencia de tokens >= 80% (por ejemplo si falta un segundo nombre menor)
        coincidencia = len(inter) / max(len(palabras_bd), len(palabras_excel))
        return coincidencia >= 0.75

    def _procesar_hoja_excel(self, df_raw: pd.DataFrame, filepath: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """Procesa una hoja individual de Excel identificando encabezados y columnas."""
        if df_raw is None or df_raw.empty:
            return None

        header_row_idx = None
        cols_norm = [self._normalizar_nombre_columna(c) for c in df_raw.columns]

        tiene_id = any(c in self.MAPEO_COLUMNAS_EXCEL and self.MAPEO_COLUMNAS_EXCEL[c] == "numero_identificacion" for c in cols_norm)

        df = df_raw
        if not tiene_id and len(df_raw) > 0:
            for r_idx in range(min(15, len(df_raw))):
                fila_valores = [self._normalizar_nombre_columna(v) for v in df_raw.iloc[r_idx].dropna()]
                if any(v in self.MAPEO_COLUMNAS_EXCEL and self.MAPEO_COLUMNAS_EXCEL[v] == "numero_identificacion" for v in fila_valores):
                    header_row_idx = r_idx + 1
                    break

            if header_row_idx is not None:
                logger.info(f"Encabezados detectados en hoja '{sheet_name}', fila {header_row_idx}")
                df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row_idx, dtype=str)

        # Normalizar columnas
        df.columns = [self._normalizar_nombre_columna(col) for col in df.columns]

        renombres = {
            col: self.MAPEO_COLUMNAS_EXCEL[col]
            for col in df.columns
            if col in self.MAPEO_COLUMNAS_EXCEL
        }
        df = df.rename(columns=renombres)

        if "numero_identificacion" not in df.columns:
            return None

        # Combinar primer_nombre y segundo_nombre si vienen separados
        if "primer_nombre" in df.columns:
            p_nom = df["primer_nombre"].fillna("").astype(str).str.strip()
            s_nom = df["segundo_nombre"].fillna("").astype(str).str.strip() if "segundo_nombre" in df.columns else ""
            noms_combinados = (p_nom + " " + s_nom).str.strip()
            if "nombres" not in df.columns or df["nombres"].isna().all():
                df["nombres"] = noms_combinados

        # Combinar primer_apellido y segundo_apellido si vienen separados
        if "primer_apellido" in df.columns:
            p_ape = df["primer_apellido"].fillna("").astype(str).str.strip()
            s_ape = df["segundo_apellido"].fillna("").astype(str).str.strip() if "segundo_apellido" in df.columns else ""
            apes_combinados = (p_ape + " " + s_ape).str.strip()
            if "apellidos" not in df.columns or df["apellidos"].isna().all():
                df["apellidos"] = apes_combinados

        # Limpiar identificaciones (ej. "CC - 1110487315" -> "1110487315")
        df["numero_identificacion"] = df["numero_identificacion"].apply(self._limpiar_numero_id)
        df = df[df["numero_identificacion"].str.len() >= 5]

        # Normalizar campos de texto
        for campo in ["nombres", "apellidos", "lugar_expedicion", "sexo"]:
            if campo in df.columns:
                df[campo] = df[campo].fillna("").astype(str).str.strip().str.upper()
                df[campo] = df[campo].replace({"NAN": "", "NONE": "", "NULL": ""})

        return df

    def cargar_excel(self, filepath: str) -> pd.DataFrame:
        """
        Carga y normaliza un archivo Excel externo leyendo todas las hojas disponibles
        (ej: 'Inscritos Primera Opción', 'Inscritos Segunda Opción').
        """
        logger.info(f"Cargando Excel externo multi-hoja: {filepath}")

        try:
            excel_file = pd.ExcelFile(filepath)
            dfs_validos: List[pd.DataFrame] = []

            for sheet_name in excel_file.sheet_names:
                try:
                    df_sheet_raw = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
                    df_procesado = self._procesar_hoja_excel(df_sheet_raw, filepath, sheet_name)
                    if df_procesado is not None and not df_procesado.empty:
                        logger.info(f"Hoja '{sheet_name}': {len(df_procesado)} registros válidos cargados")
                        dfs_validos.append(df_procesado)
                except Exception as e_sheet:
                    logger.warning(f"No se pudo procesar hoja '{sheet_name}': {e_sheet}")

            if not dfs_validos:
                raise ValueError(
                    "No se encontró ninguna columna de identificación válida (Identificación, Cédula, Documento, CC) en las hojas del archivo Excel."
                )

            df_total = pd.concat(dfs_validos, ignore_index=True)
            # Eliminar duplicados si una persona está en ambas hojas
            df_total = df_total.drop_duplicates(subset=["numero_identificacion"], keep="first")

            logger.info(f"Excel consolidado: {len(df_total)} registros únicos listos para comparar")
            return df_total

        except Exception as e:
            logger.error(f"Error cargando Excel: {e}")
            raise

    def _normalizar_para_comparacion(self, val: Any, campo: str = "") -> str:
        """
        Normaliza un valor para comparación exacta sin falsos positivos por tildes,
        espacios o formatos de fecha alternativos.
        """
        if val is None or pd.isna(val):
            return ""
        s = str(val).strip().upper()
        if s in ("NAN", "NONE", "NULL", "NAT"):
            return ""

        # Si es un campo de fecha, comparar por fecha parseada
        if "fecha" in campo or re.match(r"^\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4}", s):
            f_obj = validador.parsear_fecha(s)
            if f_obj:
                return f_obj.isoformat()

        # Quitar tildes
        repl = {"Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U", "Ü": "U"}
        for k, v in repl.items():
            s = s.replace(k, v)
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def ejecutar_comparacion(
        self,
        comparacion_id: str,
        excel_path: str,
        db: Session
    ) -> Dict[str, Any]:
        """
        Ejecuta la comparación entre BD y Excel.
        Guarda el detalle de diferencias en la BD.
        """
        inicio = time.time()
        import uuid as uuid_pkg
        logger.info(f"Ejecutando comparación {comparacion_id}")

        from app.models.comparacion import Comparacion
        from app.models.diferencia import Diferencia
        from app.models.persona import Persona

        resultado = {
            "total_registros_bd": 0,
            "total_registros_excel": 0,
            "total_coincidentes": 0,
            "total_diferentes": 0,
            "total_faltantes_bd": 0,
            "total_nuevos_bd": 0,
            "diferencias": [],
        }

        try:
            # 1. Cargar datos de BD
            personas_bd = db.query(Persona).all()
            df_bd = pd.DataFrame([{
                "numero_identificacion": self._limpiar_numero_id(p.numero_identificacion),
                "nombres": p.nombres or "",
                "apellidos": p.apellidos or "",
                "fecha_nacimiento": p.fecha_nacimiento.isoformat() if p.fecha_nacimiento else "",
                "fecha_expedicion": p.fecha_expedicion.isoformat() if p.fecha_expedicion else "",
                "lugar_expedicion": p.lugar_expedicion or "",
                "sexo": p.sexo or "",
                "pagina_numero": p.pagina_numero or 1,
                "confianza_extraccion": float(p.confianza_extraccion or 0.0),
                "estado_registro": p.estado_registro or "VALID",
            } for p in personas_bd])

            # Filtrar registros vacíos de BD
            if len(df_bd) > 0:
                df_bd = df_bd[df_bd["numero_identificacion"].str.len() >= 5]

            resultado["total_registros_bd"] = len(df_bd)

            # 2. Cargar Excel externo (todas las hojas)
            df_excel = self.cargar_excel(excel_path)
            resultado["total_registros_excel"] = len(df_excel)

            # 3. Analizar diferencias
            diferencias_a_guardar = []
            comp_uuid = uuid_pkg.UUID(str(comparacion_id))

            if len(df_bd) > 0:
                ids_bd = set(df_bd["numero_identificacion"].tolist())
                ids_excel = set(df_excel["numero_identificacion"].tolist())

                ids_faltantes = ids_excel - ids_bd      # En Excel, no en BD
                ids_nuevos = ids_bd - ids_excel          # En BD, no en Excel
                ids_comunes = ids_bd & ids_excel         # En ambos

                resultado["total_faltantes_bd"] = len(ids_faltantes)
                resultado["total_nuevos_bd"] = len(ids_nuevos)

                # Diferencias en registros faltantes en BD
                for id_faltante in ids_faltantes:
                    row_excel = df_excel[df_excel["numero_identificacion"] == id_faltante].iloc[0]
                    diferencias_a_guardar.append(Diferencia(
                        comparacion_id=comp_uuid,
                        numero_identificacion=id_faltante,
                        campo="registro_completo",
                        valor_bd=None,
                        valor_excel=str(row_excel.to_dict()),
                        tipo_diferencia="faltante_bd",
                    ))

                # Registros nuevos en BD
                for id_nuevo in ids_nuevos:
                    row_bd = df_bd[df_bd["numero_identificacion"] == id_nuevo].iloc[0]
                    diferencias_a_guardar.append(Diferencia(
                        comparacion_id=comp_uuid,
                        numero_identificacion=id_nuevo,
                        campo="registro_completo",
                        valor_bd=str(row_bd.to_dict()),
                        valor_excel=None,
                        tipo_diferencia="nuevo_bd",
                    ))

                # Comparar campos de registros comunes con normalización inteligente
                total_iguales = 0
                total_diferentes = 0

                tiene_col_apellidos = "apellidos" in df_excel.columns and not df_excel["apellidos"].isna().all()

                for id_comun in ids_comunes:
                    row_bd = df_bd[df_bd["numero_identificacion"] == id_comun].iloc[0]
                    row_excel = df_excel[df_excel["numero_identificacion"] == id_comun].iloc[0]

                    tiene_diferencias = False

                    # Comparación especial de Nombres y Apellidos
                    if "nombres" in row_excel:
                        if not tiene_col_apellidos:
                            # Caso SENA / Planilla con una sola columna 'Nombre' que incluye apellidos
                            noms_coinciden = self._son_nombres_equivalentes(
                                nombres_bd=str(row_bd.get("nombres") or ""),
                                apellidos_bd=str(row_bd.get("apellidos") or ""),
                                nombres_excel=str(row_excel.get("nombres") or ""),
                                apellidos_excel=""
                            )
                            if not noms_coinciden:
                                tiene_diferencias = True
                                nombre_bd_completo = f"{row_bd.get('nombres', '')} {row_bd.get('apellidos', '')}".strip()
                                diferencias_a_guardar.append(Diferencia(
                                    comparacion_id=comp_uuid,
                                    numero_identificacion=id_comun,
                                    campo="nombres",
                                    valor_bd=nombre_bd_completo,
                                    valor_excel=str(row_excel.get("nombres") or ""),
                                    tipo_diferencia="diferente",
                                ))
                        else:
                            # Columnas separadas
                            val_bd_norm = self._normalizar_para_comparacion(row_bd.get("nombres"), campo="nombres")
                            val_excel_norm = self._normalizar_para_comparacion(row_excel.get("nombres"), campo="nombres")
                            if val_bd_norm != val_excel_norm:
                                tiene_diferencias = True
                                diferencias_a_guardar.append(Diferencia(
                                    comparacion_id=comp_uuid,
                                    numero_identificacion=id_comun,
                                    campo="nombres",
                                    valor_bd=str(row_bd.get("nombres") or "") or None,
                                    valor_excel=str(row_excel.get("nombres") or "") or None,
                                    tipo_diferencia="diferente",
                                ))

                    # Comparar los demás campos
                    campos_resto = ["apellidos", "fecha_nacimiento", "fecha_expedicion", "lugar_expedicion", "sexo"]
                    if not tiene_col_apellidos:
                        campos_resto.remove("apellidos")

                    for campo in campos_resto:
                        if campo in row_excel and str(row_excel.get(campo) or "").strip():
                            val_bd_norm = self._normalizar_para_comparacion(row_bd.get(campo), campo=campo)
                            val_excel_norm = self._normalizar_para_comparacion(row_excel.get(campo), campo=campo)

                            if val_bd_norm != val_excel_norm:
                                tiene_diferencias = True
                                diferencias_a_guardar.append(Diferencia(
                                    comparacion_id=comp_uuid,
                                    numero_identificacion=id_comun,
                                    campo=campo,
                                    valor_bd=str(row_bd.get(campo) or "") or None,
                                    valor_excel=str(row_excel.get(campo) or "") or None,
                                    tipo_diferencia="diferente",
                                ))

                    if tiene_diferencias:
                        total_diferentes += 1
                    else:
                        total_iguales += 1

                resultado["total_coincidentes"] = total_iguales
                resultado["total_diferentes"] = total_diferentes

            else:
                # BD vacía: todo el Excel son registros faltantes
                resultado["total_faltantes_bd"] = len(df_excel)
                for _, row in df_excel.iterrows():
                    diferencias_a_guardar.append(Diferencia(
                        comparacion_id=comp_uuid,
                        numero_identificacion=row["numero_identificacion"],
                        campo="registro_completo",
                        valor_bd=None,
                        valor_excel=str(row.to_dict()),
                        tipo_diferencia="faltante_bd",
                    ))


            # 4. Guardar diferencias en BD (por lotes)
            BATCH_SIZE = 500
            for i in range(0, len(diferencias_a_guardar), BATCH_SIZE):
                lote = diferencias_a_guardar[i:i + BATCH_SIZE]
                db.add_all(lote)
                db.commit()

            # 5. Actualizar comparación
            tiempo_ms = int((time.time() - inicio) * 1000)
            comparacion = db.query(Comparacion).filter(Comparacion.id == comp_uuid).first()
            if comparacion:
                comparacion.total_registros_bd = resultado["total_registros_bd"]
                comparacion.total_registros_excel = resultado["total_registros_excel"]
                comparacion.total_coincidentes = resultado["total_coincidentes"]
                comparacion.total_diferentes = resultado["total_diferentes"]
                comparacion.total_faltantes_bd = resultado["total_faltantes_bd"]
                comparacion.total_nuevos_bd = resultado["total_nuevos_bd"]
                comparacion.estado = "completado"
                comparacion.fecha_ejecucion = datetime.utcnow()
                comparacion.tiempo_procesamiento_ms = tiempo_ms
                db.commit()

            logger.info(
                f"Comparación completada en {tiempo_ms}ms. "
                f"Iguales: {resultado['total_coincidentes']}, "
                f"Diferentes: {resultado['total_diferentes']}, "
                f"Faltantes: {resultado['total_faltantes_bd']}, "
                f"Nuevos: {resultado['total_nuevos_bd']}"
            )

            return resultado

        except Exception as e:
            logger.error(f"Error en comparación: {e}")
            try:
                comparacion = db.query(Comparacion).filter(Comparacion.id == comparacion_id).first()
                if comparacion:
                    comparacion.estado = "error"
                    comparacion.mensaje_error = str(e)[:500]
                    db.commit()
            except Exception:
                pass
            raise

    def generar_reporte_xlsx(self, comparacion_id: str, db: Session) -> str:
        """
        Genera un reporte XLSX profesional con el mismo estándar corporativo del módulo de exportación.
        Incluye:
        - Hoja 1: Resumen Ejecutivo con KPIs, porcentajes de concordancia e instrucciones
        - Hoja 2: Auditoría General con todos los registros cotejados (Frente OCR vs Excel Oficial)
        - Hoja 3: Campos con Diferencias (celdas resaltadas en amarillo para cotejo visual inmediato)
        - Hoja 4: Faltantes en BD (personas en Excel que no fueron detectadas por OCR)
        - Hoja 5: Sobrantes en BD (personas en BD que no estaban en la planilla oficial)
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from app.models.comparacion import Comparacion
        from app.models.diferencia import Diferencia
        from app.models.persona import Persona
        from app.config import settings

        comparacion = db.query(Comparacion).filter(Comparacion.id == comparacion_id).first()
        if not comparacion:
            raise ValueError("Comparación no encontrada")

        # 1. Cargar datos del Excel original si existe
        df_excel = pd.DataFrame()
        excel_map: Dict[str, Dict[str, Any]] = {}
        if comparacion.ruta_archivo and Path(comparacion.ruta_archivo).exists():
            try:
                df_excel = self.cargar_excel(comparacion.ruta_archivo)
                if not df_excel.empty and "numero_identificacion" in df_excel.columns:
                    for _, r in df_excel.iterrows():
                        num_id = str(r["numero_identificacion"]).strip()
                        if num_id:
                            excel_map[num_id] = r.to_dict()
            except Exception as e:
                logger.warning(f"No se pudo cargar el archivo Excel original: {e}")

        # 2. Cargar personas de la BD
        personas_bd = db.query(Persona).all()
        bd_map: Dict[str, Persona] = {
            str(p.numero_identificacion).strip(): p for p in personas_bd if p.numero_identificacion
        }

        # 3. Cargar diferencias calculadas
        diferencias = db.query(Diferencia).filter(
            Diferencia.comparacion_id == comparacion_id
        ).all()
        difs_por_id: Dict[str, List[Diferencia]] = {}
        for d in diferencias:
            difs_por_id.setdefault(str(d.numero_identificacion).strip(), []).append(d)

        # 4. Construir universo completo de registros
        todos_los_ids = sorted(list(set(excel_map.keys()) | set(bd_map.keys())))

        registros_auditoria = []
        for id_num in todos_los_ids:
            en_bd = id_num in bd_map
            en_excel = id_num in excel_map
            p = bd_map.get(id_num)
            r_ex = excel_map.get(id_num, {})
            difs = difs_por_id.get(id_num, [])

            difs_campos = {d.campo: d for d in difs if d.tipo_diferencia == "diferente"}

            if en_bd and not en_excel:
                estado = "SOBRANTE EN BD"
                detalle = "Detectado por OCR pero no registrado en la planilla Excel"
            elif en_excel and not en_bd:
                estado = "FALTANTE EN BD"
                detalle = "En planilla Excel pero NO detectado por OCR en los documentos PDF"
            elif difs_campos:
                estado = "DIFERENCIA DE DATOS"
                detalles_l = []
                for c, d in difs_campos.items():
                    detalles_l.append(f"{c.upper()}: OCR='{d.valor_bd}' vs Excel='{d.valor_excel}'")
                detalle = " | ".join(detalles_l)
            else:
                estado = "COINCIDENTE"
                detalle = "Datos 100% coincidentes"

            # Formatear fechas
            f_nac_bd = p.fecha_nacimiento.strftime("%d/%m/%Y") if (p and p.fecha_nacimiento) else ""
            f_nac_ex = str(r_ex.get("fecha_nacimiento") or "")
            if f_nac_ex and f_nac_ex != "None":
                try:
                    f_nac_ex = datetime.strptime(f_nac_ex[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    pass
            else:
                f_nac_ex = ""

            f_exp_bd = p.fecha_expedicion.strftime("%d/%m/%Y") if (p and p.fecha_expedicion) else ""
            f_exp_ex = str(r_ex.get("fecha_expedicion") or "")
            if f_exp_ex and f_exp_ex != "None":
                try:
                    f_exp_ex = datetime.strptime(f_exp_ex[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    pass
            else:
                f_exp_ex = ""

            nombres_ex = str(r_ex.get("nombres") or "")
            if nombres_ex == "None": nombres_ex = ""
            apellidos_ex = str(r_ex.get("apellidos") or "")
            if apellidos_ex == "None": apellidos_ex = ""
            lugar_ex = str(r_ex.get("lugar_expedicion") or "")
            if lugar_ex == "None": lugar_ex = ""
            sexo_ex = str(r_ex.get("sexo") or "")
            if sexo_ex == "None": sexo_ex = ""

            conf_ocr = f"{float(p.confianza_extraccion or 0):.1f}%" if p else "—"
            estado_ocr = ("REVISAR" if p.requiere_revision else "VÁLIDO") if p else "NO DETECTADO"

            registros_auditoria.append({
                "identificacion": id_num,
                "estado": estado,
                "nombres_bd": p.nombres if p and p.nombres else "",
                "nombres_excel": nombres_ex,
                "apellidos_bd": p.apellidos if p and p.apellidos else "",
                "apellidos_excel": apellidos_ex,
                "f_nac_bd": f_nac_bd,
                "f_nac_excel": f_nac_ex,
                "f_exp_bd": f_exp_bd,
                "f_exp_excel": f_exp_ex,
                "lugar_bd": p.lugar_expedicion if p and p.lugar_expedicion else "",
                "lugar_excel": lugar_ex,
                "sexo_bd": p.sexo if p and p.sexo else "",
                "sexo_excel": sexo_ex,
                "confianza_ocr": conf_ocr,
                "estado_ocr": estado_ocr,
                "detalle_discrepancias": detalle,
                "difs_campos": difs_campos
            })

        # Ordenar: primero Diferencias, luego Faltantes, luego Sobrantes, luego Coincidentes
        orden_estado = {"DIFERENCIA DE DATOS": 1, "FALTANTE EN BD": 2, "SOBRANTE EN BD": 3, "COINCIDENTE": 4}
        registros_auditoria.sort(key=lambda r: orden_estado.get(r["estado"], 99))

        # 5. Crear Workbook OpenPyXL y aplicar estilos corporativos
        wb = openpyxl.Workbook()

        # Estilos compartidos
        COLOR_HEADER = "1E3A5F"      # Azul oscuro
        COLOR_SUBHEADER = "2E86AB"   # Azul medio
        COLOR_ALTROW = "F8FAFC"      # Blanco grisáceo sutil
        COLOR_OK = "D4EDDA"          # Verde suave coincidente
        COLOR_DIF = "FFF3CD"         # Amarillo suave diferencias
        COLOR_FAL = "F8D7DA"         # Rojo suave faltante en BD
        COLOR_SOB = "E8F4FD"         # Azul suave sobrante en BD

        fuente_header = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        fuente_datos = Font(name="Calibri", size=9.5)
        fuente_diff = Font(name="Calibri", bold=True, size=9.5, color="856404")
        relleno_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        relleno_ok = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
        relleno_dif = PatternFill(start_color=COLOR_DIF, end_color=COLOR_DIF, fill_type="solid")
        relleno_fal = PatternFill(start_color=COLOR_FAL, end_color=COLOR_FAL, fill_type="solid")
        relleno_sob = PatternFill(start_color=COLOR_SOB, end_color=COLOR_SOB, fill_type="solid")
        relleno_alt = PatternFill(start_color=COLOR_ALTROW, end_color=COLOR_ALTROW, fill_type="solid")
        borde = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )

        headers = [
            "N° Identificación", "Estado Auditoría",
            "Nombres (OCR / BD)", "Nombres (Excel Oficial)",
            "Apellidos (OCR / BD)", "Apellidos (Excel Oficial)",
            "F. Nacimiento (OCR)", "F. Nacimiento (Excel)",
            "F. Expedición (OCR)", "F. Expedición (Excel)",
            "Lugar Expedición (OCR)", "Lugar Expedición (Excel)",
            "Sexo (OCR)", "Sexo (Excel)",
            "Confianza OCR", "Estado OCR",
            "Detalle de Discrepancias / Observaciones"
        ]

        # ── HOJA 1: RESUMEN EJECUTIVO ──────────────────────────────
        ws_res = wb.active
        ws_res.title = "Resumen Ejecutivo"
        ws_res.views.sheetView[0].showGridLines = True

        ws_res.merge_cells("A1:G1")
        ws_res["A1"] = "SISTEMA OCR - AUDITORÍA Y COMPARACIÓN DE DATOS"
        ws_res["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        ws_res["A1"].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws_res["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_res.row_dimensions[1].height = 35

        ws_res.merge_cells("A2:G2")
        ws_res["A2"] = "Reporte de Control de Calidad y Discrepancias vs Listado Oficial"
        ws_res["A2"].font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        ws_res["A2"].fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type="solid")
        ws_res["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws_res.row_dimensions[2].height = 20

        # Bloque 1: Metadatos
        ws_res.merge_cells("A4:G4")
        ws_res["A4"] = "INFORMACIÓN DEL PROCESAMIENTO"
        ws_res["A4"].font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
        ws_res["A4"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws_res["A4"].alignment = Alignment(horizontal="left", vertical="center")
        ws_res.row_dimensions[4].height = 24

        meta_rows = [
            ("Archivo Oficial Comparado:", comparacion.nombre_original),
            ("Fecha de Ejecución:", comparacion.fecha_ejecucion.strftime("%d/%m/%Y %H:%M") if comparacion.fecha_ejecucion else "—"),
            ("Total Personas en Base de Datos (OCR):", comparacion.total_registros_bd),
            ("Total Registros en Archivo Excel:", comparacion.total_registros_excel),
        ]
        for idx, (etq, val) in enumerate(meta_rows, start=5):
            ws_res.cell(row=idx, column=1, value=etq).font = Font(name="Calibri", bold=True, size=10)
            ws_res.cell(row=idx, column=2, value=val).font = Font(name="Calibri", size=10)
            ws_res.row_dimensions[idx].height = 20

        # Bloque 2: KPIs
        ws_res.merge_cells("A10:G10")
        ws_res["A10"] = "MÉTRICAS DE CONTROL DE CALIDAD"
        ws_res["A10"].font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
        ws_res["A10"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws_res["A10"].alignment = Alignment(horizontal="left", vertical="center")
        ws_res.row_dimensions[10].height = 24

        kpi_headers = ["Métrica", "Cantidad", "% Sobre Total Excel", "Estado / Significado"]
        for col_idx, kh in enumerate(kpi_headers, start=1):
            c = ws_res.cell(row=11, column=col_idx, value=kh)
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = relleno_header
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws_res.row_dimensions[11].height = 24

        tot_ex = max(1, comparacion.total_registros_excel or 1)
        pct_ok = round((comparacion.total_coincidentes / tot_ex) * 100.0, 1)
        pct_dif = round((comparacion.total_diferentes / tot_ex) * 100.0, 1)
        pct_fal = round((comparacion.total_faltantes_bd / tot_ex) * 100.0, 1)

        kpi_data = [
            ("Registros Coincidentes (100% OK)", comparacion.total_coincidentes, f"{pct_ok}%", "Datos idénticos entre OCR y Excel Oficial", COLOR_OK),
            ("Registros con Diferencias de Datos", comparacion.total_diferentes, f"{pct_dif}%", "Cédula encontrada pero con datos discordantes (requiere revisión)", COLOR_DIF),
            ("Registros Faltantes en BD (No Detectados)", comparacion.total_faltantes_bd, f"{pct_fal}%", "En planilla Excel pero NO detectados en los documentos PDF", COLOR_FAL),
            ("Registros Sobrantes en BD (Extras)", comparacion.total_nuevos_bd, "N/A", "Detectados por OCR pero no registrados en la planilla Excel", COLOR_SOB),
            ("Tasa Global de Concordancia", f"{pct_ok}%", "100%", "Porcentaje de coincidencia exacta respecto a la planilla oficial", "D1E7DD"),
        ]

        for idx, (metrica, cant, pct, desc, col_fill) in enumerate(kpi_data, start=12):
            ws_res.cell(row=idx, column=1, value=metrica).font = Font(name="Calibri", bold=True, size=10)
            ws_res.cell(row=idx, column=2, value=cant).font = Font(name="Calibri", bold=True, size=10)
            ws_res.cell(row=idx, column=2).alignment = Alignment(horizontal="center")
            ws_res.cell(row=idx, column=3, value=pct).font = Font(name="Calibri", bold=True, size=10)
            ws_res.cell(row=idx, column=3).alignment = Alignment(horizontal="center")
            ws_res.cell(row=idx, column=4, value=desc).font = Font(name="Calibri", size=9.5)
            ws_res.cell(row=idx, column=1).fill = PatternFill(start_color=col_fill, end_color=col_fill, fill_type="solid")
            for c_i in range(1, 5):
                ws_res.cell(row=idx, column=c_i).border = borde
            ws_res.row_dimensions[idx].height = 22

        # Bloque 3: Instrucciones
        ws_res.merge_cells("A18:G18")
        ws_res["A18"] = "GUÍA DE AUDITORÍA Y CORRECCIÓN"
        ws_res["A18"].font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
        ws_res["A18"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        ws_res["A18"].alignment = Alignment(horizontal="left", vertical="center")
        ws_res.row_dimensions[18].height = 24

        instrucciones = [
            "1. Revise la hoja 'Campos con Diferencias' para observar los valores en amarillo que no coinciden con la planilla oficial.",
            "2. Revise la hoja 'Faltantes en BD' para identificar qué personas de la lista oficial no fueron detectadas en el PDF.",
            "3. Revise la hoja 'Sobrantes en BD' para auditar documentos detectados que no pertenecen a la lista oficial.",
            "4. En el sistema web (/comparacion o /personas), utilice el botón 'Corregir' para actualizar cualquier dato erróneo."
        ]
        for idx, inst in enumerate(instrucciones, start=19):
            ws_res.merge_cells(f"A{idx}:G{idx}")
            c_inst = ws_res[f"A{idx}"]
            c_inst.value = inst
            c_inst.font = Font(name="Calibri", italic=True, size=9.5, color="334155")
            ws_res.row_dimensions[idx].height = 20

        for col_i in range(1, 8):
            ws_res.column_dimensions[get_column_letter(col_i)].width = 30

        # Función auxiliar para escribir hojas de datos estructurados
        def _escribir_hoja_datos(ws, items, solo_diferencias=False):
            ws.views.sheetView[0].showGridLines = True
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=1, column=col_idx, value=h)
                c.font = fuente_header
                c.fill = relleno_header
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = borde
            ws.row_dimensions[1].height = 30

            for row_idx, r in enumerate(items, start=2):
                st = r["estado"]
                es_alt = (row_idx % 2 == 0)

                if st == "COINCIDENTE":
                    r_fill = relleno_ok
                elif st == "DIFERENCIA DE DATOS":
                    r_fill = relleno_dif
                elif st == "FALTANTE EN BD":
                    r_fill = relleno_fal
                else:
                    r_fill = relleno_sob

                difs_c = r.get("difs_campos", {})

                row_vals = [
                    (r["identificacion"], "left", False),
                    (r["estado"], "center", False),
                    (r["nombres_bd"], "left", "nombres" in difs_c),
                    (r["nombres_excel"], "left", "nombres" in difs_c),
                    (r["apellidos_bd"], "left", "apellidos" in difs_c),
                    (r["apellidos_excel"], "left", "apellidos" in difs_c),
                    (r["f_nac_bd"], "center", "fecha_nacimiento" in difs_c),
                    (r["f_nac_excel"], "center", "fecha_nacimiento" in difs_c),
                    (r["f_exp_bd"], "center", "fecha_expedicion" in difs_c),
                    (r["f_exp_excel"], "center", "fecha_expedicion" in difs_c),
                    (r["lugar_bd"], "left", "lugar_expedicion" in difs_c),
                    (r["lugar_excel"], "left", "lugar_expedicion" in difs_c),
                    (r["sexo_bd"], "center", "sexo" in difs_c),
                    (r["sexo_excel"], "center", "sexo" in difs_c),
                    (r["confianza_ocr"], "center", False),
                    (r["estado_ocr"], "center", False),
                    (r["detalle_discrepancias"], "left", False),
                ]

                for col_idx, (val, alig, es_diferente) in enumerate(row_vals, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.alignment = Alignment(horizontal=alig, vertical="center")
                    c.border = borde

                    if es_diferente:
                        c.fill = relleno_dif
                        c.font = fuente_diff
                    elif col_idx == 2:
                        c.fill = r_fill
                        c.font = Font(name="Calibri", bold=True, size=9.5)
                    elif es_alt:
                        c.fill = relleno_alt
                        c.font = fuente_datos
                    else:
                        c.font = fuente_datos

                ws.row_dimensions[row_idx].height = 20

            ws.freeze_panes = "C2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(items) + 1}"

            # Autoajustar ancho de columnas
            for col_i, h in enumerate(headers, start=1):
                max_len = len(h)
                for r_i in range(2, min(len(items) + 2, 50)):
                    v = str(ws.cell(row=r_i, column=col_i).value or "")
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[get_column_letter(col_i)].width = min(45, max(max_len + 3, 12))

        # ── HOJA 2: AUDITORÍA GENERAL ──────────────────────────────
        ws_aud = wb.create_sheet(title="Auditoría General")
        _escribir_hoja_datos(ws_aud, registros_auditoria)

        # ── HOJA 3: CAMPOS CON DIFERENCIAS ─────────────────────────
        registros_dif = [r for r in registros_auditoria if r["estado"] == "DIFERENCIA DE DATOS"]
        if registros_dif:
            ws_dif = wb.create_sheet(title="Campos con Diferencias")
            _escribir_hoja_datos(ws_dif, registros_dif, solo_diferencias=True)

        # ── HOJA 4: FALTANTES EN BD ────────────────────────────────
        registros_fal = [r for r in registros_auditoria if r["estado"] == "FALTANTE EN BD"]
        if registros_fal:
            ws_fal = wb.create_sheet(title="Faltantes en BD")
            _escribir_hoja_datos(ws_fal, registros_fal)

        # ── HOJA 5: SOBRANTES EN BD ────────────────────────────────
        registros_sob = [r for r in registros_auditoria if r["estado"] == "SOBRANTE EN BD"]
        if registros_sob:
            ws_sob = wb.create_sheet(title="Sobrantes en BD")
            _escribir_hoja_datos(ws_sob, registros_sob)

        # Guardar archivo generado
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre = f"reporte_auditoria_comparacion_{timestamp}.xlsx"
        ruta = settings.export_path / nombre
        wb.save(str(ruta))

        logger.info(f"Reporte profesional de comparación generado exitosamente: {ruta}")
        return str(ruta)


comparacion_service = ComparacionService()
