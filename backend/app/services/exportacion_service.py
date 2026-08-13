"""
Servicio de exportación a Excel (XLSX)
Usa Pandas + OpenPyXL para generar reportes formateados
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session

from app.utils.logger import app_logger as logger
from app.config import settings


class ExportacionService:
    """Servicio de exportación de datos a archivos XLSX con formato profesional"""

    COLUMNAS = {
        "numero_identificacion": "Número Identificación",
        "nombres": "Nombres",
        "apellidos": "Apellidos",
        "fecha_nacimiento": "Fecha Nacimiento",
        "fecha_expedicion": "Fecha Expedición",
        "lugar_expedicion": "Lugar Expedición",
        "sexo": "Sexo",
        "confianza_extraccion": "Confianza OCR (%)",
        "requiere_revision": "Requiere Revisión",
        "fecha_registro": "Fecha Registro",
    }

    # Colores corporativos
    COLOR_HEADER = "1E3A5F"      # Azul oscuro
    COLOR_SUBHEADER = "2E86AB"   # Azul medio
    COLOR_ALTROW = "EBF4FA"      # Azul muy claro
    COLOR_REVISION = "FFF3CD"    # Amarillo suave
    COLOR_OK = "D4EDDA"          # Verde suave

    def exportar_personas(self, db: Session, filtros: Optional[dict] = None) -> str:
        """
        Exporta todas las personas de la BD a un archivo XLSX formateado.
        
        Returns:
            Ruta del archivo generado
        """
        logger.info("Iniciando exportación de personas a XLSX")

        from app.models.persona import Persona

        # Consulta con filtros opcionales
        query = db.query(Persona)
        if filtros:
            if filtros.get("requiere_revision") is not None:
                query = query.filter(Persona.requiere_revision == filtros["requiere_revision"])

        personas = query.order_by(Persona.fecha_registro.desc()).all()
        logger.info(f"Exportando {len(personas)} registros")

        # Convertir a DataFrame
        datos = []
        for p in personas:
            datos.append({
                "numero_identificacion": p.numero_identificacion,
                "nombres": p.nombres or "",
                "apellidos": p.apellidos or "",
                "fecha_nacimiento": p.fecha_nacimiento.isoformat() if p.fecha_nacimiento else "",
                "fecha_expedicion": p.fecha_expedicion.isoformat() if p.fecha_expedicion else "",
                "lugar_expedicion": p.lugar_expedicion or "",
                "sexo": p.sexo or "",
                "confianza_extraccion": float(p.confianza_extraccion or 0),
                "requiere_revision": "SÍ" if p.requiere_revision else "NO",
                "fecha_registro": p.fecha_registro.strftime("%d/%m/%Y %H:%M") if p.fecha_registro else "",
            })

        df = pd.DataFrame(datos, columns=list(self.COLUMNAS.keys()))
        df.columns = list(self.COLUMNAS.values())

        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"personas_ocr_{timestamp}.xlsx"
        ruta_archivo = settings.export_path / nombre_archivo

        # Guardar con Pandas primero
        df.to_excel(str(ruta_archivo), index=False, engine="openpyxl")

        # Aplicar formato profesional con OpenPyXL
        self._aplicar_formato_excel(str(ruta_archivo), len(datos))

        logger.info(f"Archivo exportado: {ruta_archivo}")
        return str(ruta_archivo)

    def _aplicar_formato_excel(self, filepath: str, total_filas: int):
        """Aplica formato visual profesional al archivo Excel"""
        wb = load_workbook(filepath)
        ws = wb.active
        ws.title = "Personas Registradas"

        # Fuentes
        fuente_header = Font(
            name="Calibri", bold=True, size=11, color="FFFFFF"
        )
        fuente_titulo = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
        fuente_datos = Font(name="Calibri", size=10)

        # Rellenos
        relleno_header = PatternFill(
            start_color=self.COLOR_HEADER,
            end_color=self.COLOR_HEADER,
            fill_type="solid"
        )
        relleno_altrow = PatternFill(
            start_color=self.COLOR_ALTROW,
            end_color=self.COLOR_ALTROW,
            fill_type="solid"
        )
        relleno_revision = PatternFill(
            start_color=self.COLOR_REVISION,
            end_color=self.COLOR_REVISION,
            fill_type="solid"
        )
        relleno_ok = PatternFill(
            start_color=self.COLOR_OK,
            end_color=self.COLOR_OK,
            fill_type="solid"
        )

        # Borde
        borde = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        # Insertar fila de título
        ws.insert_rows(1)
        titulo_cell = ws["A1"]
        titulo_cell.value = "SISTEMA OCR - REGISTRO DE PERSONAS"
        titulo_cell.font = fuente_titulo
        titulo_cell.fill = PatternFill(
            start_color=self.COLOR_SUBHEADER,
            end_color=self.COLOR_SUBHEADER,
            fill_type="solid"
        )
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(f"A1:{get_column_letter(len(self.COLUMNAS))}1")
        ws.row_dimensions[1].height = 35

        # Insertar fila de metadatos
        ws.insert_rows(2)
        meta_cell = ws["A2"]
        meta_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total registros: {total_filas}"
        meta_cell.font = Font(name="Calibri", italic=True, size=9, color="555555")
        ws.merge_cells(f"A2:{get_column_letter(len(self.COLUMNAS))}2")
        ws.row_dimensions[2].height = 20

        # Formatear fila de encabezados (ahora fila 3)
        for col, header in enumerate(self.COLUMNAS.values(), start=1):
            cell = ws.cell(row=3, column=col)
            cell.font = fuente_header
            cell.fill = relleno_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde
        ws.row_dimensions[3].height = 30

        # Ajustar ancho de columnas y formatear datos
        anchos = [20, 25, 25, 18, 18, 25, 10, 16, 18, 20]
        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        # Formatear filas de datos
        for row_num in range(4, total_filas + 4):
            es_altrow = (row_num % 2 == 0)
            requiere_revision = ws.cell(row=row_num, column=9).value == "SÍ"

            for col_num in range(1, len(self.COLUMNAS) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = fuente_datos
                cell.border = borde
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Colorear según estado de revisión
                if requiere_revision:
                    cell.fill = relleno_revision
                elif es_altrow:
                    cell.fill = relleno_altrow
                else:
                    cell.fill = relleno_ok

            ws.row_dimensions[row_num].height = 18

        # Inmovilizar paneles
        ws.freeze_panes = "A4"

        # Filtros automáticos
        ws.auto_filter.ref = f"A3:{get_column_letter(len(self.COLUMNAS))}{total_filas + 3}"

        # Hoja de resumen
        ws_resumen = wb.create_sheet("Resumen")
        self._crear_hoja_resumen(ws_resumen, total_filas)

        wb.save(filepath)
        logger.info("Formato Excel aplicado correctamente")

    def _crear_hoja_resumen(self, ws, total_filas: int):
        """Crea una hoja de resumen estadístico"""
        ws.title = "Resumen"
        ws["A1"] = "RESUMEN DE EXPORTACIÓN"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Total de personas exportadas"
        ws["B3"] = total_filas
        ws["A4"] = "Fecha de exportación"
        ws["B4"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    def exportar_reporte_diferencias(self, db: Session, comparacion_id: str) -> str:
        """
        Exporta el reporte estructurado de diferencias a Excel XLSX.
        11 Columnas: Número Documento | Página | Tipo Documento | Campo | Valor OCR | Valor Excel | Diferencia | Confianza OCR | Estado | Motor OCR | Motivo
        """
        from app.models.diferencia import Diferencia
        from app.models.persona import Persona

        difs = db.query(Diferencia).filter(Diferencia.comparacion_id == comparacion_id).all()
        datos = []
        for d in difs:
            p = db.query(Persona).filter(Persona.numero_identificacion == d.numero_identificacion).first()
            pag = p.pagina_numero if p and p.pagina_numero else 1
            tipo_doc = p.tipo_documento if p and p.tipo_documento else "CEDULA_CIUDADANIA"
            motor = p.motor_ocr if p and p.motor_ocr else "google_document_ai"
            conf = float(p.confianza_extraccion or 0) if p else 0.0
            est = p.estado_registro if p and p.estado_registro else "REVIEW_REQUIRED"

            motivo = "Los valores no coinciden"
            if d.tipo_diferencia == "faltante_bd":
                motivo = "Registro presente en Excel pero ausente en BD"
            elif d.tipo_diferencia == "nuevo_bd":
                motivo = "Registro presente en BD pero ausente en Excel"

            datos.append({
                "Número Documento": d.numero_identificacion,
                "Página": pag,
                "Tipo Documento": tipo_doc,
                "Campo": d.campo,
                "Valor OCR": d.valor_bd or "",
                "Valor Excel": d.valor_excel or "",
                "Diferencia": d.tipo_diferencia.upper(),
                "Confianza OCR": conf,
                "Estado": est,
                "Motor OCR": motor,
                "Motivo": motivo,
            })

        if not datos:
            datos.append({
                "Número Documento": "N/A", "Página": "-", "Tipo Documento": "-", "Campo": "-",
                "Valor OCR": "-", "Valor Excel": "-", "Diferencia": "SIN DIFERENCIAS",
                "Confianza OCR": 100.0, "Estado": "VALID", "Motor OCR": "-", "Motivo": "Todas las coincidencia son perfectas"
            })

        df = pd.DataFrame(datos)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_diferencias_{timestamp}.xlsx"
        ruta_archivo = settings.export_path / nombre_archivo
        df.to_excel(str(ruta_archivo), index=False, engine="openpyxl")
        return str(ruta_archivo)


exportacion_service = ExportacionService()
