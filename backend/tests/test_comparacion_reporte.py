import pytest
import openpyxl
import pandas as pd
from pathlib import Path
from unittest.mock import MagicMock
from datetime import datetime

from app.services.comparacion_service import comparacion_service


def test_generar_reporte_xlsx_formato(tmp_path):
    # 1. Crear archivo Excel temporal
    excel_path = tmp_path / "test_planilla.xlsx"
    df_excel_test = pd.DataFrame([
        {
            "identificacion": "1117493336",
            "nombres": "LIDA YASMIN",
            "apellidos": "ALDANA BOHORQUEZ",
            "fecha_nacimiento": "1987-05-08",
            "fecha_expedicion": "2005-05-19",
            "lugar_expedicion": "FLORENCIA",
            "sexo": "F"
        },
        {
            "identificacion": "7556032",
            "nombres": "PABLO",
            "apellidos": "RODRIGUEZ RENGIFO",
            "fecha_nacimiento": "1969-01-02",
            "fecha_expedicion": "1987-04-12",
            "lugar_expedicion": "ARMENIA",
            "sexo": "M"
        },
        {
            "identificacion": "1000000001",
            "nombres": "CARLOS",
            "apellidos": "PEREZ LOPEZ",
            "fecha_nacimiento": "1995-10-10",
            "fecha_expedicion": "2013-10-10",
            "lugar_expedicion": "BOGOTA",
            "sexo": "M"
        }
    ])
    df_excel_test.to_excel(str(excel_path), index=False)

    # Mock de DB, Comparacion, Diferencia, Persona
    mock_db = MagicMock()
    mock_comp = MagicMock()
    mock_comp.id = "comp-123"
    mock_comp.nombre_original = "test_planilla.xlsx"
    mock_comp.ruta_archivo = str(excel_path)
    mock_comp.fecha_ejecucion = datetime(2026, 9, 3, 10, 0)
    mock_comp.total_registros_bd = 3
    mock_comp.total_registros_excel = 3
    mock_comp.total_coincidentes = 1
    mock_comp.total_diferentes = 1
    mock_comp.total_faltantes_bd = 1
    mock_comp.total_nuevos_bd = 1

    # Mock de personas en BD
    # p1: Coincidente
    mock_p1 = MagicMock()
    mock_p1.numero_identificacion = "1117493336"
    mock_p1.nombres = "LIDA YASMIN"
    mock_p1.apellidos = "ALDANA BOHORQUEZ"
    mock_p1.fecha_nacimiento = datetime(1987, 5, 8).date()
    mock_p1.fecha_expedicion = datetime(2005, 5, 19).date()
    mock_p1.lugar_expedicion = "FLORENCIA"
    mock_p1.sexo = "F"
    mock_p1.confianza_extraccion = 94.8
    mock_p1.requiere_revision = False

    # p2: Con diferencia en apellidos
    mock_p2 = MagicMock()
    mock_p2.numero_identificacion = "7556032"
    mock_p2.nombres = "PABLO"
    mock_p2.apellidos = "RODRIGUEZ"
    mock_p2.fecha_nacimiento = datetime(1969, 1, 2).date()
    mock_p2.fecha_expedicion = datetime(1987, 4, 12).date()
    mock_p2.lugar_expedicion = "ARMENIA"
    mock_p2.sexo = "M"
    mock_p2.confianza_extraccion = 90.0
    mock_p2.requiere_revision = True

    # p3: Sobrante en BD (no está en el Excel)
    mock_p3 = MagicMock()
    mock_p3.numero_identificacion = "9999999999"
    mock_p3.nombres = "EXTRA"
    mock_p3.apellidos = "SOBRANTE"
    mock_p3.fecha_nacimiento = datetime(2000, 1, 1).date()
    mock_p3.fecha_expedicion = datetime(2018, 1, 1).date()
    mock_p3.lugar_expedicion = "CALI"
    mock_p3.sexo = "M"
    mock_p3.confianza_extraccion = 85.0
    mock_p3.requiere_revision = False

    # Diferencia para p2
    mock_dif = MagicMock()
    mock_dif.numero_identificacion = "7556032"
    mock_dif.campo = "apellidos"
    mock_dif.valor_bd = "RODRIGUEZ"
    mock_dif.valor_excel = "RODRIGUEZ RENGIFO"
    mock_dif.tipo_diferencia = "diferente"

    def mock_query(model):
        m = MagicMock()
        name = getattr(model, '__name__', str(model))
        if "Comparacion" in name:
            m.filter.return_value.first.return_value = mock_comp
        elif "Persona" in name:
            m.all.return_value = [mock_p1, mock_p2, mock_p3]
        elif "Diferencia" in name:
            m.filter.return_value.all.return_value = [mock_dif]
        return m

    mock_db.query = mock_query

    # Ejecutar generación
    ruta = comparacion_service.generar_reporte_xlsx("comp-123", mock_db)
    assert Path(ruta).exists()
    assert ruta.endswith(".xlsx")

    # Validar estructura y hojas con OpenPyXL
    wb = openpyxl.load_workbook(ruta)
    assert "Resumen Ejecutivo" in wb.sheetnames
    assert "Auditoría General" in wb.sheetnames
    assert "Campos con Diferencias" in wb.sheetnames
    assert "Faltantes en BD" in wb.sheetnames
    assert "Sobrantes en BD" in wb.sheetnames

    # Validar banner en Resumen Ejecutivo
    ws_res = wb["Resumen Ejecutivo"]
    assert "SISTEMA OCR - AUDITORÍA Y COMPARACIÓN DE DATOS" in str(ws_res["A1"].value)

    # Validar registros en Auditoría General
    ws_aud = wb["Auditoría General"]
    assert ws_aud["A1"].value == "N° Identificación"
    assert ws_aud["B1"].value == "Estado Auditoría"

    # Validar hoja de diferencias
    ws_dif = wb["Campos con Diferencias"]
    assert ws_dif["A2"].value == "7556032"

    # Validar hoja de faltantes
    ws_fal = wb["Faltantes en BD"]
    assert ws_fal["A2"].value == "1000000001"

    # Validar hoja de sobrantes
    ws_sob = wb["Sobrantes en BD"]
    assert ws_sob["A2"].value == "9999999999"

    wb.close()
    print("Test passed: 5 hojas generadas con precisión y validadas.")
